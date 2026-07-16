import os, csv, hashlib, smtplib, json
from email.mime.text import MIMEText
from datetime import datetime
from openpyxl import load_workbook

INTAKE_FILE = "new_tenancies.xlsx"
PHOTO_DIR = "movein_photos"
REGISTER = "tenancy_register.csv"
OUTBOX = "outbox"
DRY_RUN = True  # set False and fill SMTP settings to send real emails

def validate(row):
    errors = []
    if not row["tenant_email"] or "@" not in row["tenant_email"]:
        errors.append("invalid tenant email")
    if not row["landlord_email"] or "@" not in row["landlord_email"]:
        errors.append("invalid landlord email")
    try:
        if float(row["deposit_eur"]) <= 0:
            errors.append("deposit must be positive")
    except (ValueError, TypeError):
        errors.append("deposit not a number")
    if not row["address"]:
        errors.append("missing address")
    try:
        datetime.strptime(str(row["start_date"])[:10], "%Y-%m-%d")
    except ValueError:
        errors.append("bad start date")
    return errors

def hash_photos(tenancy_id):
    folder = os.path.join(PHOTO_DIR, tenancy_id)
    hashes = {}
    if not os.path.isdir(folder):
        return hashes
    for name in sorted(os.listdir(folder)):
        with open(os.path.join(folder, name), "rb") as f:
            hashes[name] = hashlib.sha256(f.read()).hexdigest()
    return hashes

def send_email(to, subject, body):
    msg = MIMEText(body)
    msg["Subject"] = subject
    msg["To"] = to
    msg["From"] = "noreply@depositguard.ie"
    if DRY_RUN:
        os.makedirs(OUTBOX, exist_ok=True)
        fname = subject.lower().replace(" ", "_")[:40] + "_" + to.split("@")[0] + ".eml"
        with open(os.path.join(OUTBOX, fname), "w") as f:
            f.write(msg.as_string())
        return
    with smtplib.SMTP("smtp.gmail.com", 587) as s:
        s.starttls()
        s.login(os.environ["SMTP_USER"], os.environ["SMTP_PASS"])
        s.send_message(msg)

def run():
    wb = load_workbook(INTAKE_FILE)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    processed_col = headers.index("processed") + 1
    processed, rejected = 0, 0

    exists = os.path.isfile(REGISTER)
    reg = open(REGISTER, "a", newline="")
    writer = csv.writer(reg)
    if not exists:
        writer.writerow(["tenancy_id", "address", "deposit_eur", "tenant_email",
                         "landlord_email", "start_date", "photo_hashes", "status", "processed_at"])

    for row_num, row_cells in enumerate(ws.iter_rows(min_row=2), start=2):
        row = dict(zip(headers, [c.value for c in row_cells]))
        if row.get("processed") == "yes":
            continue
        tenancy_id = str(row["tenancy_id"])
        errors = validate(row)
        if errors:
            print(f"[REJECTED] {tenancy_id}: {', '.join(errors)}")
            send_email(str(row["landlord_email"] or "admin@depositguard.ie"),
                       f"DepositGuard intake rejected {tenancy_id}",
                       "Your tenancy intake was rejected:\n- " + "\n- ".join(errors))
            ws.cell(row=row_num, column=processed_col, value="yes")
            rejected += 1
            continue

        hashes = hash_photos(tenancy_id)
        writer.writerow([tenancy_id, row["address"], row["deposit_eur"], row["tenant_email"],
                         row["landlord_email"], str(row["start_date"])[:10],
                         json.dumps(hashes), "escrow_created", datetime.now().isoformat(timespec="seconds")])

        body = (f"Escrow created for {row['address']}.\n"
                f"Deposit: EUR {row['deposit_eur']}\n"
                f"Move-in photos hashed: {len(hashes)}\n"
                f"Tenancy ID: {tenancy_id}")
        send_email(row["tenant_email"], f"DepositGuard escrow ready {tenancy_id}", body)
        send_email(row["landlord_email"], f"DepositGuard escrow ready {tenancy_id}", body)
        ws.cell(row=row_num, column=processed_col, value="yes")
        print(f"[OK] {tenancy_id} -> escrow created, {len(hashes)} photos hashed")
        processed += 1

    reg.close()
    wb.save(INTAKE_FILE)
    print(f"\nDone. {processed} processed, {rejected} rejected.")

if __name__ == "__main__":
    run()
